from openpyxl import Workbook

#Input file names
files=[]
while True:
	files.append(input('File name?\n'))
	if files[-1]=='':
		break

#Setup workbook
wb=Workbook()
ws=wb.active

j=1
items=['Time,s','T,hrs','CH4','CO','CO2','T,oven','T,f','T,out','T,in']
for item in items:
	ws.cell(row=1,column=j,value=item)
	j=j+1

#Define resample function
def resample(file_name,offset):
	n=0
	f=open(file_name,mode='r')
	for line in f:
		i=0
		l=line
		l=l.split('\t')
		if l[0]=='X_Value':
			n=1
			continue
		if n==1:
			t0=float(l[0])
		if n>=1 and n%24==0:
			t1=float(l[0])
			t=(t1-t0)/3600
			l.insert(1,t)
			for item in l:
				ws.cell(row=n//24+1,column=i+offset,value=float(item))
				i=i+1
		if n>=1:
			n=n+1
	return len(l)+1
	f.close()

k=0
for file in files:
        try:
        	print(k)
        	k=k+resample(file,k+1)
        except:
        	pass
wb.save(files[0]+'_resampled.xlsx')
